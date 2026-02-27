from __future__ import annotations

import csv
import io
import json
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation
from urllib.parse import parse_qs, urlencode, urlparse, urlunparse
from urllib.request import urlopen

from django.utils.text import slugify

from apps.paineis.models import Dataset

DATE_FORMATS = [
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%Y/%m/%d",
    "%d/%m/%y",
]

TRUE_VALUES = {"1", "true", "t", "sim", "s", "yes"}
FALSE_VALUES = {"0", "false", "f", "nao", "não", "n", "no"}

SENSITIVE_HINTS = {
    "cpf",
    "cnpj",
    "rg",
    "telefone",
    "celular",
    "email",
    "e-mail",
    "endereco",
    "endereço",
    "logradouro",
    "nascimento",
}


def _clean_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_headers(raw_headers: list[str]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for i, header in enumerate(raw_headers, start=1):
        base = (header or "").strip() or f"coluna_{i}"
        base = base.replace("\n", " ").strip()
        key = slugify(base).replace("-", "_") or f"coluna_{i}"
        if key in seen:
            seen[key] += 1
            key = f"{key}_{seen[key]}"
        else:
            seen[key] = 1
        headers.append(key)
    return headers


def _parse_date(value: str):
    text = (value or "").strip()
    if not text:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_number(value: str):
    text = (value or "").strip()
    if not text:
        return None

    clean = text.replace("R$", "").replace(" ", "")
    if "," in clean and "." in clean:
        if clean.rfind(",") > clean.rfind("."):
            clean = clean.replace(".", "").replace(",", ".")
        else:
            clean = clean.replace(",", "")
    elif "," in clean:
        clean = clean.replace(".", "").replace(",", ".")

    try:
        return Decimal(clean)
    except (InvalidOperation, ValueError):
        return None


def _parse_bool(value: str):
    text = (value or "").strip().lower()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return None


def _infer_column_type(values: list[str]) -> str:
    non_empty = [v for v in values if v != ""]
    if not non_empty:
        return "TEXTO"

    bool_hits = sum(1 for v in non_empty if _parse_bool(v) is not None)
    date_hits = sum(1 for v in non_empty if _parse_date(v) is not None)
    number_hits = sum(1 for v in non_empty if _parse_number(v) is not None)

    total = len(non_empty)
    required = total if total <= 5 else max(1, int(total * 0.8))
    if bool_hits == total:
        return "BOOLEANO"
    if date_hits >= required:
        return "DATA"
    if number_hits >= required:
        return "NUMERO"
    return "TEXTO"


def _detect_sensitive(header: str) -> bool:
    check = (header or "").strip().lower().replace("_", " ")
    return any(hint in check for hint in SENSITIVE_HINTS)


def _read_csv_rows(raw: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = raw.decode("utf-8-sig", errors="ignore")
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        raise ValueError("Arquivo CSV está vazio.")

    sample = "\n".join(lines[:5])
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = ";" if ";" in sample else ","

    reader = csv.reader(io.StringIO("\n".join(lines)), delimiter=delimiter)
    all_rows = list(reader)
    if not all_rows:
        raise ValueError("Arquivo CSV sem linhas válidas.")

    headers = _normalize_headers([_clean_cell(v) for v in all_rows[0]])
    rows: list[dict[str, str]] = []
    for row_values in all_rows[1:]:
        row_dict: dict[str, str] = {}
        for idx, header in enumerate(headers):
            row_dict[header] = _clean_cell(row_values[idx]) if idx < len(row_values) else ""
        rows.append(row_dict)

    return headers, rows


def _read_xlsx_rows(raw: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise ValueError("Leitura XLSX indisponível. Instale openpyxl para habilitar.") from exc

    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb.active

    header_row = None
    rows: list[dict[str, str]] = []

    for idx, values in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [_clean_cell(v) for v in values]
        if idx == 1:
            header_row = cells
            continue
        if not header_row:
            continue

        if not any(cells):
            continue

        headers = _normalize_headers(header_row)
        row_dict: dict[str, str] = {}
        for cidx, header in enumerate(headers):
            row_dict[header] = cells[cidx] if cidx < len(cells) else ""
        rows.append(row_dict)

    if not header_row:
        raise ValueError("Planilha XLSX sem cabeçalho na primeira linha.")

    return _normalize_headers(header_row), rows


def _google_sheet_csv_url(url: str) -> str:
    raw = (url or "").strip()
    if not raw:
        raise ValueError("URL de Google Sheets não informada.")

    parsed = urlparse(raw)
    if "docs.google.com" not in parsed.netloc.lower():
        return raw

    path = parsed.path or ""
    if "/export" in path and "format=csv" in parsed.query:
        return raw

    query = parse_qs(parsed.query)
    gid = (query.get("gid") or [""])[0]

    if "/edit" in path:
        path = path.split("/edit", 1)[0] + "/export"

    params = {"format": "csv"}
    if gid:
        params["gid"] = gid

    return urlunparse((parsed.scheme, parsed.netloc, path, "", urlencode(params), ""))


def _fetch_google_sheet(url: str) -> bytes:
    final_url = _google_sheet_csv_url(url)
    with urlopen(final_url, timeout=20) as resp:
        return resp.read()


def _build_schema_and_profile(headers: list[str], rows: list[dict[str, str]]) -> tuple[list[dict], dict]:
    schema: list[dict] = []
    nulls_by_column: dict[str, int] = {}
    duplicates_count = 0
    seen_rows: set[tuple] = set()
    numeric_stats: dict[str, dict] = {}

    for header in headers:
        values = [r.get(header, "") for r in rows]
        col_type = _infer_column_type(values[:2000])
        role = "MEDIDA" if col_type == "NUMERO" else "DIMENSAO"
        sample = next((v for v in values if v), "")

        schema.append(
            {
                "name": header,
                "type": col_type,
                "role": role,
                "sensitive": _detect_sensitive(header),
                "sample": sample[:120],
            }
        )
        nulls_by_column[header] = sum(1 for v in values if not v)

        if col_type == "NUMERO":
            parsed = [_parse_number(v) for v in values]
            nums = [n for n in parsed if n is not None]
            if nums:
                numeric_stats[header] = {
                    "min": str(min(nums)),
                    "max": str(max(nums)),
                    "sum": str(sum(nums)),
                }

    for row in rows[:20000]:
        marker = tuple(row.get(h, "") for h in headers)
        if marker in seen_rows:
            duplicates_count += 1
        else:
            seen_rows.add(marker)

    warnings: list[str] = []
    if duplicates_count:
        warnings.append(f"Foram detectadas {duplicates_count} linhas duplicadas na amostra.")

    type_counter = Counter(item["type"] for item in schema)

    profile = {
        "row_count": len(rows),
        "column_count": len(headers),
        "nulls_by_column": nulls_by_column,
        "duplicate_rows_sample": duplicates_count,
        "types_summary": dict(type_counter),
        "numeric_stats": numeric_stats,
        "warnings": warnings,
    }

    return schema, profile


def _build_processed_csv(headers: list[str], rows: list[dict[str, str]]) -> bytes:
    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=headers, delimiter=";")
    writer.writeheader()
    for row in rows:
        writer.writerow({h: row.get(h, "") for h in headers})
    text = out.getvalue()
    return ("\ufeff" + text).encode("utf-8")


def ingest_dataset_bytes(
    raw: bytes,
    fonte: str,
    *,
    filename: str = "",
    google_sheet_url: str = "",
) -> dict:
    source = (fonte or "").strip().upper()

    if source == Dataset.Fonte.GOOGLE_SHEETS:
        raw = _fetch_google_sheet(google_sheet_url)
        headers, rows = _read_csv_rows(raw)
    elif source == Dataset.Fonte.CSV:
        headers, rows = _read_csv_rows(raw)
    elif source == Dataset.Fonte.XLSX:
        headers, rows = _read_xlsx_rows(raw)
    elif source in {Dataset.Fonte.PDF, Dataset.Fonte.DOCX}:
        raise ValueError(
            "Extração de PDF/DOCX ainda é condicional no MVP. Exporte para CSV/XLSX para ingestão segura."
        )
    else:
        raise ValueError("Fonte de dados não suportada para ingestão.")

    if not headers:
        raise ValueError("Não foi possível identificar colunas no arquivo enviado.")

    schema, profile = _build_schema_and_profile(headers, rows)
    processed_csv = _build_processed_csv(headers, rows)

    return {
        "source": source,
        "filename": filename,
        "headers": headers,
        "rows": rows,
        "preview_rows": rows[:40],
        "schema": schema,
        "profile": profile,
        "processed_csv_bytes": processed_csv,
        "warnings": profile.get("warnings", []),
    }


def data_dictionary_csv(schema: list[dict]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["coluna", "tipo", "papel", "sensivel", "amostra"])
    for col in schema:
        writer.writerow(
            [
                col.get("name", ""),
                col.get("type", ""),
                col.get("role", ""),
                "sim" if col.get("sensitive") else "não",
                col.get("sample", ""),
            ]
        )
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def profile_json_bytes(profile: dict) -> bytes:
    return json.dumps(profile or {}, ensure_ascii=False, indent=2).encode("utf-8")
